from pathlib import Path
from datetime import datetime

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from src.analytics.relative_age import build_relative_age_summary
from src.reports.quality import build_quality_report
from src.version import __version__

HEADER_FILL = PatternFill("solid", fgColor="111827")
HEADER_FONT = Font(color="FFFFFF", bold=True)


def _format_sheet(ws) -> None:
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for column_cells in ws.columns:
        letter = get_column_letter(column_cells[0].column)
        max_length = max(len(str(cell.value or "")) for cell in column_cells[:300])
        ws.column_dimensions[letter].width = min(max(max_length + 2, 10), 34)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top")


def export_workbook(df: pd.DataFrame, output_path: Path, source: str, input_type: str, dataset_name: str, competition: str) -> Path:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    quality = build_quality_report(df)
    if "team_code" in df.columns:
        teams = df.groupby(["team", "team_code"], dropna=False).size().reset_index(name="players").sort_values("team")
    else:
        teams = df.groupby("team").size().reset_index(name="players").sort_values("team")
    rae = build_relative_age_summary(df)
    metadata = pd.DataFrame([
        ["software", "SportsLab-Analytics"], ["version", __version__],
        ["organization", "SportsLabResearch"],
        ["generated_at", datetime.now().isoformat(timespec="seconds")],
        ["dataset_name", dataset_name], ["competition", competition],
        ["input_type", input_type], ["source", source],
        ["players", len(df)], ["teams", int(df["team"].nunique())],
    ], columns=["field", "value"])
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Players", index=False)
        teams.to_excel(writer, sheet_name="Teams", index=False)
        rae.to_excel(writer, sheet_name="Relative_Age", index=False)
        quality.to_excel(writer, sheet_name="Data_Quality", index=False)
        metadata.to_excel(writer, sheet_name="Metadata", index=False)
        for worksheet in writer.book.worksheets:
            _format_sheet(worksheet)
        if "birth_date" in df.columns:
            birth_column = list(df.columns).index("birth_date") + 1
            for cell in writer.book["Players"][get_column_letter(birth_column)][1:]:
                cell.number_format = "dd/mm/yyyy"
    return output_path
